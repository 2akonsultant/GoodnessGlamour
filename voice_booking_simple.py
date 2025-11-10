"""
Simplified Voice Booking Assistant - No pandas/numpy dependencies
Lightweight version for easy installation
"""

import os
import json
import uuid
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, asdict
from flask import Flask, request, Response, jsonify, render_template_string
from twilio.twiml.voice_response import VoiceResponse
from twilio.rest import Client
import qrcode
import base64
from io import BytesIO
import openpyxl
import sqlite3
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configuration - Twilio Credentials
TWILIO_ACCOUNT_SID = os.getenv('TWILIO_ACCOUNT_SID', 'ACd8941e7d6933a9e031879bc28d7af7e8')
TWILIO_AUTH_TOKEN = os.getenv('TWILIO_AUTH_TOKEN', '815e90983ed99b02e52943cc14602d56')
TWILIO_PHONE_NUMBER = os.getenv('TWILIO_PHONE_NUMBER', '+917019035686')
TWILIO_RECOVERY_CODE = os.getenv('TWILIO_RECOVERY_CODE', 'J4VXS81L35XTR3ZK54V8TX7A')
TWILIO_SECRET_KEY = os.getenv('TWILIO_SECRET_KEY', 'dUytv2kejZ0E3jPTgUyNCXV92zzfulJw')
WEBHOOK_BASE_URL = os.getenv('WEBHOOK_BASE_URL', 'https://your-domain.com')

# Simulation Mode (for testing without actual phone calls)
SIMULATION_MODE = os.getenv('SIMULATION_MODE', 'true').lower() == 'true'

# Initialize Twilio client with enhanced error handling
try:
    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
    logger.info("✅ Twilio client initialized successfully")
except Exception as e:
    logger.error(f"❌ Failed to initialize Twilio client: {e}")
    client = None

def verify_twilio_account():
    """Verify Twilio account credentials and get account info"""
    if not client:
        return False
    
    try:
        # Get account information
        account = client.api.accounts(TWILIO_ACCOUNT_SID).fetch()
        logger.info(f"✅ Twilio Account Verified: {account.friendly_name}")
        logger.info(f"📊 Account Status: {account.status}")
        logger.info(f"💰 Account Balance: {account.balance}")
        
        # Get available phone numbers
        incoming_phone_numbers = client.incoming_phone_numbers.list(limit=5)
        logger.info(f"📞 Available Phone Numbers: {len(incoming_phone_numbers)}")
        for number in incoming_phone_numbers:
            logger.info(f"   📱 {number.phone_number} - {number.friendly_name}")
        
        return True
    except Exception as e:
        logger.error(f"❌ Twilio account verification failed: {e}")
        return False

# Salon Information
SALON_INFO = {
    "name": "Goodness Glamour Salon",
    "phone": "9036626642",
    "email": "2akonsultant@gmail.com",
    "hours": "Monday - Sunday, 9:00 AM - 8:00 PM",
    "service_type": "Doorstep beauty services"
}

# Services and Pricing
SERVICES = {
    "women": {
        "haircut": {"name": "Women's Haircut & Styling", "price": "₹400-1,200", "duration": "60 minutes"},
        "coloring": {"name": "Hair Coloring & Highlights", "price": "₹1,200-3,500", "duration": "120 minutes"},
        "treatment": {"name": "Hair Treatment & Spa", "price": "₹600-2,000", "duration": "90 minutes"},
        "bridal": {"name": "Bridal & Party Styling", "price": "₹800-2,500", "duration": "90 minutes"},
        "blowdry": {"name": "Professional Blowdry", "price": "₹250-600", "duration": "45 minutes"},
        "hairwash": {"name": "Hair Wash & Styling", "price": "₹200-450", "duration": "30 minutes"},
        "consultation": {"name": "Hair Consultation", "price": "₹150-300", "duration": "30 minutes"}
    },
    "kids": {
        "haircut": {"name": "Kids Haircuts", "price": "₹150-500", "duration": "30 minutes"},
        "party": {"name": "Party Styling", "price": "₹200-600", "duration": "45 minutes"},
        "hairwash": {"name": "Kids Hair Wash", "price": "₹100-300", "duration": "20 minutes"},
        "braiding": {"name": "Creative Braiding", "price": "₹150-400", "duration": "30 minutes"}
    }
}

@dataclass
class BookingData:
    customer_name: str = ""
    phone: str = ""
    service: str = ""
    date: str = ""
    time: str = ""
    address: str = ""
    notes: str = ""

@dataclass
class ConversationState:
    session_id: str
    customer_name: Optional[str] = None
    phone: Optional[str] = None
    current_step: str = "greeting"
    booking_data: BookingData = None
    conversation_history: List[Dict] = None
    
    def __post_init__(self):
        if self.booking_data is None:
            self.booking_data = BookingData()
        if self.conversation_history is None:
            self.conversation_history = []

class SimpleVoiceAssistant:
    """
    Simplified AI Voice Assistant without heavy dependencies
    """
    
    def __init__(self):
        self.active_sessions: Dict[str, ConversationState] = {}
        self._ensure_directories()
    
    def _ensure_directories(self):
        """Ensure data directories exist"""
        os.makedirs("data", exist_ok=True)
        os.makedirs("logs", exist_ok=True)
    
    def get_conversation_response(self, user_input: str, session_id: str) -> str:
        """Generate AI response based on conversation state and user input"""
        session = self.active_sessions.get(session_id)
        if not session:
            return "I'm sorry, I'm having trouble with this call. Please try calling again."
        
        # Simple rule-based responses for voice calls
        current_step = session.current_step
        user_input_lower = user_input.lower().strip()
        
        if current_step == "greeting":
            if any(word in user_input_lower for word in ["book", "appointment", "schedule", "service"]):
                session.current_step = "get_name"
                return "Great! I'd love to help you book an appointment. What's your name?"
            else:
                return "Hello! Welcome to Goodness Glamour Salon. I'm your AI assistant. How can I help you today?"
        
        elif current_step == "get_name":
            # Extract name from input
            name = self._extract_name(user_input)
            if name:
                session.customer_name = name
                session.current_step = "get_service"
                return f"Nice to meet you, {name}! Which service would you like? We do haircuts, coloring, treatments, and bridal styling."
            else:
                return "I didn't catch your name. Could you please tell me your name?"
        
        elif current_step == "get_service":
            service = self._identify_service(user_input)
            if service:
                session.booking_data.service = service
                session.current_step = "get_date"
                service_info = self._get_service_info(service)
                return f"Perfect! {service_info['name']} is {service_info['price']}. What date works for you?"
            else:
                return "Which service interests you? We do haircuts, coloring, treatments, and bridal styling."
        
        elif current_step == "get_date":
            date = self._parse_date(user_input)
            if date:
                session.booking_data.date = date
                session.current_step = "get_time"
                return f"Great! For {date}, what time would work? We're available 9 AM to 8 PM."
            else:
                return "What date would you like? You can say tomorrow, or a specific date."
        
        elif current_step == "get_time":
            time = self._parse_time(user_input)
            if time:
                session.booking_data.time = time
                session.current_step = "get_address"
                return "Perfect! What's your address for our doorstep service?"
            else:
                return "What time works for you? We're available 9 AM to 8 PM."
        
        elif current_step == "get_address":
            address = user_input.strip()
            if len(address) > 10:  # Basic validation
                session.booking_data.address = address
                session.current_step = "confirm_booking"
                return self._generate_confirmation_message(session)
            else:
                return "Could you please provide your complete address for our doorstep service?"
        
        elif current_step == "confirm_booking":
            if any(word in user_input_lower for word in ["yes", "confirm", "correct", "book"]):
                session.current_step = "booking_complete"
                return self._complete_booking(session)
            elif any(word in user_input_lower for word in ["no", "change", "wrong"]):
                session.current_step = "get_name"  # Restart
                return "No problem! Let's start over. What's your name?"
            else:
                return "Please say 'yes' to confirm or 'no' to make changes."
        
        elif current_step == "booking_complete":
            return "Thank you for choosing Goodness Glamour! Have a wonderful day!"
        
        return "I'm not sure how to help with that. Could you please repeat?"

    def _extract_name(self, text: str) -> Optional[str]:
        """Extract name from user input"""
        text_lower = text.lower()
        
        # Common patterns for name extraction
        if "my name is" in text_lower:
            # Extract name after "my name is"
            parts = text.split("my name is", 1)
            if len(parts) > 1:
                name = parts[1].strip().split()[0]
                return name.capitalize()
        elif "i'm" in text_lower or "i am" in text_lower:
            # Extract name after "i'm" or "i am"
            parts = text.split()
            for i, word in enumerate(parts):
                if word.lower() in ["i'm", "i"] and i + 1 < len(parts) and parts[i + 1].lower() == "am":
                    if i + 2 < len(parts):
                        return parts[i + 2].capitalize()
                elif word.lower() == "i'm" and i + 1 < len(parts):
                    return parts[i + 1].capitalize()
        
        # Fallback: take first word
        words = text.split()
        if len(words) >= 1:
            return words[0].capitalize()
        return None

    def _identify_service(self, text: str) -> Optional[str]:
        """Identify service from user input"""
        text_lower = text.lower()
        
        service_mapping = {
            "haircut": ["haircut", "cut", "trim"],
            "coloring": ["color", "coloring", "highlight", "dye"],
            "treatment": ["treatment", "spa", "keratin", "therapy"],
            "bridal": ["bridal", "wedding", "party", "styling"],
            "blowdry": ["blowdry", "blow dry", "style"],
            "hairwash": ["wash", "shampoo", "clean"],
            "consultation": ["consultation", "advice", "consult"]
        }
        
        for service, keywords in service_mapping.items():
            if any(keyword in text_lower for keyword in keywords):
                return service
        
        return None

    def _get_service_info(self, service: str) -> Dict:
        """Get service information"""
        for category in SERVICES.values():
            if service in category:
                return category[service]
        return {"name": "Service", "price": "Contact us", "duration": "Varies"}

    def _parse_date(self, text: str) -> Optional[str]:
        """Parse date from user input"""
        text_lower = text.lower()
        
        if "tomorrow" in text_lower:
            tomorrow = datetime.now() + timedelta(days=1)
            return tomorrow.strftime("%A, %B %d")
        elif "today" in text_lower:
            return datetime.now().strftime("%A, %B %d")
        else:
            return text.strip()

    def _parse_time(self, text: str) -> Optional[str]:
        """Parse time from user input"""
        text_lower = text.lower()
        
        if any(word in text_lower for word in ["morning", "am"]):
            return "10 AM"
        elif any(word in text_lower for word in ["afternoon", "pm"]):
            return "2 PM"
        elif any(word in text_lower for word in ["evening"]):
            return "6 PM"
        else:
            return text.strip()

    def _generate_confirmation_message(self, session: ConversationState) -> str:
        """Generate booking confirmation message"""
        booking = session.booking_data
        return f"""Let me confirm your booking:
Name: {session.customer_name}
Service: {booking.service}
Date: {booking.date}
Time: {booking.time}
Address: {booking.address}
Does this look correct?"""

    def _complete_booking(self, session: ConversationState) -> str:
        """Complete the booking and return confirmation"""
        booking_id = f"BG{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Save booking
        self._save_booking(session, booking_id)
        
        # Send SMS confirmation
        self._send_sms_confirmation(session, booking_id)
        
        return f"""Perfect! Your booking is confirmed. Booking ID: {booking_id}. 
You'll receive an SMS confirmation shortly. 
Thank you for choosing Goodness Glamour Salon!"""

    def _save_booking(self, session: ConversationState, booking_id: str):
        """Save booking to Excel and database"""
        booking_data = {
            "booking_id": booking_id,
            "customer_name": session.customer_name,
            "phone": session.phone,
            "service": session.booking_data.service,
            "date": session.booking_data.date,
            "time": session.booking_data.time,
            "address": session.booking_data.address,
            "status": "confirmed",
            "created_at": datetime.now().isoformat(),
            "source": "voice_call"
        }
        
        # Save to Excel
        self._save_to_excel(booking_data)
        
        # Save to SQLite
        self._save_to_database(booking_data)
        
        logger.info(f"Booking saved: {booking_data}")

    def _save_to_excel(self, booking_data: Dict):
        """Save booking to Excel file using openpyxl"""
        try:
            file_path = "data/bookings.xlsx"
            
            # Try to load existing workbook
            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
            else:
                # Create new workbook
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Bookings"
                
                # Add headers
                headers = ['Booking ID', 'Customer Name', 'Phone', 'Service', 'Date', 'Time', 'Address', 'Status', 'Created At', 'Source']
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
            # Add new booking
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=booking_data['booking_id'])
            ws.cell(row=row, column=2, value=booking_data['customer_name'])
            ws.cell(row=row, column=3, value=booking_data['phone'])
            ws.cell(row=row, column=4, value=booking_data['service'])
            ws.cell(row=row, column=5, value=booking_data['date'])
            ws.cell(row=row, column=6, value=booking_data['time'])
            ws.cell(row=row, column=7, value=booking_data['address'])
            ws.cell(row=row, column=8, value=booking_data['status'])
            ws.cell(row=row, column=9, value=booking_data['created_at'])
            ws.cell(row=row, column=10, value=booking_data['source'])
            
            wb.save(file_path)
            logger.info(f"Booking saved to Excel: {booking_data['booking_id']}")
            
        except Exception as e:
            logger.error(f"Error saving to Excel: {e}")

    def _save_to_database(self, booking_data: Dict):
        """Save booking to SQLite database"""
        try:
            conn = sqlite3.connect("data/salon_bookings.db")
            cursor = conn.cursor()
            
            # Create table if not exists
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS voice_bookings (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    booking_id TEXT UNIQUE NOT NULL,
                    customer_name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    service TEXT NOT NULL,
                    date TEXT NOT NULL,
                    time TEXT NOT NULL,
                    address TEXT NOT NULL,
                    status TEXT DEFAULT 'confirmed',
                    source TEXT DEFAULT 'voice_call',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT
                )
            ''')
            
            # Insert booking
            cursor.execute('''
                INSERT OR REPLACE INTO voice_bookings 
                (booking_id, customer_name, phone, service, date, time, address, status, source, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                booking_data['booking_id'],
                booking_data['customer_name'],
                booking_data['phone'],
                booking_data['service'],
                booking_data['date'],
                booking_data['time'],
                booking_data['address'],
                booking_data['status'],
                booking_data['source'],
                booking_data.get('notes', '')
            ))
            
            conn.commit()
            conn.close()
            logger.info(f"Booking saved to database: {booking_data['booking_id']}")
            
        except Exception as e:
            logger.error(f"Error saving to database: {e}")

    def _send_sms_confirmation(self, session: ConversationState, booking_id: str):
        """Send SMS confirmation to customer with enhanced error handling"""
        try:
            # Create customer confirmation message
            customer_message = f"""🎉 Goodness Glamour Salon - Booking Confirmed!

📋 Booking ID: {booking_id}
👤 Customer: {session.customer_name}
💇‍♀️ Service: {session.booking_data.service}
📅 Date: {session.booking_data.date}
⏰ Time: {session.booking_data.time}
📍 Address: {session.booking_data.address}

🚗 We'll be at your doorstep at the scheduled time.
📞 Contact: 9036626642 for any queries.

Thank you for choosing Goodness Glamour! 💐"""
            
            # Create salon notification message
            salon_message = f"""🔔 New Voice Booking Alert!

📋 Booking ID: {booking_id}
👤 Customer: {session.customer_name}
📞 Phone: {session.phone}
💇‍♀️ Service: {session.booking_data.service}
📅 Date: {session.booking_data.date}
⏰ Time: {session.booking_data.time}
📍 Address: {session.booking_data.address}
🔗 Source: AI Voice Assistant

Booked at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"""
            
            # Send SMS to customer
            customer_sms = client.messages.create(
                body=customer_message,
                from_=TWILIO_PHONE_NUMBER,
                to=session.phone
            )
            logger.info(f"✅ Customer SMS sent to {session.phone} - SID: {customer_sms.sid}")
            
            # Send notification to salon (optional - using salon's own number)
            try:
                salon_sms = client.messages.create(
                    body=salon_message,
                    from_=TWILIO_PHONE_NUMBER,
                    to=TWILIO_PHONE_NUMBER  # Send to salon's own number
                )
                logger.info(f"✅ Salon notification sent - SID: {salon_sms.sid}")
            except Exception as salon_error:
                logger.warning(f"⚠️ Salon notification failed: {salon_error}")
            
        except Exception as e:
            logger.error(f"❌ Failed to send SMS: {e}")
            # Log the error details for debugging
            if "21659" in str(e):
                logger.error("🔧 SMS Error: Phone number format issue. Check TWILIO_PHONE_NUMBER in .env")
            elif "21211" in str(e):
                logger.error("🔧 SMS Error: Invalid phone number format")
            elif "20003" in str(e):
                logger.error("🔧 SMS Error: Authentication failed. Check Twilio credentials")
            else:
                logger.error(f"🔧 SMS Error Details: {e}")

    def start_session(self, call_sid: str, phone_number: str) -> ConversationState:
        """Start a new conversation session"""
        session = ConversationState(
            session_id=call_sid,
            phone=phone_number
        )
        self.active_sessions[call_sid] = session
        return session

    def end_session(self, call_sid: str):
        """End conversation session"""
        if call_sid in self.active_sessions:
            del self.active_sessions[call_sid]

    def generate_qr_code(self, service_id: str = None, source: str = "website") -> dict:
        """Generate QR code data and image"""
        try:
            # Create QR code data
            qr_data = {
                "url": f"{WEBHOOK_BASE_URL}/qr/voice-booking",
                "service_id": service_id,
                "source": source,
                "timestamp": datetime.now().isoformat(),
                "type": "voice_booking"
            }
            
            # Generate QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(json.dumps(qr_data))
            qr.make(fit=True)
            
            # Create QR code image
            img = qr.make_image(fill_color="black", back_color="white")
            
            # Convert to base64
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            qr_id = f"qr_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            return {
                "qr_id": qr_id,
                "qr_data": qr_data,
                "qr_image": img_base64,
                "qr_url": f"{WEBHOOK_BASE_URL}/qr/voice-booking"
            }
            
        except Exception as e:
            logger.error(f"Error generating QR code: {e}")
            return None

    def trigger_voice_call(self, phone_number: str, source: str = "qr_code") -> dict:
        """Trigger voice call to customer"""
        
        # Check if we're in simulation mode
        if SIMULATION_MODE:
            return self._simulate_voice_call(phone_number, source)
        
        try:
            # Make outbound call to customer
            call = client.calls.create(
                url=f'{WEBHOOK_BASE_URL}/voice/incoming',
                to=phone_number,
                from_=TWILIO_PHONE_NUMBER,
                method='POST',
                status_callback=f'{WEBHOOK_BASE_URL}/voice/status',
                status_callback_method='POST'
            )
            
            logger.info(f"Voice call triggered to {phone_number} from {source}, SID: {call.sid}")
            
            return {
                "success": True,
                "call_sid": call.sid,
                "message": "Voice call initiated successfully",
                "phone_number": phone_number,
                "source": source
            }
            
        except Exception as e:
            logger.error(f"Error triggering voice call: {e}")
            logger.info("🔄 Falling back to simulation mode...")
            return self._simulate_voice_call(phone_number, source)
    
    def _simulate_voice_call(self, phone_number: str, source: str) -> dict:
        """Simulate voice call for testing"""
        import uuid
        import time
        
        call_sid = f"sim_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        
        logger.info(f"🎭 SIMULATING voice call to {phone_number} from {source}")
        logger.info(f"   Call SID: {call_sid}")
        logger.info(f"   📞 This is a simulation - no actual call will be made")
        
        # Simulate the conversation flow
        self._simulate_conversation_flow(call_sid, phone_number)
        
        return {
            "success": True,
            "call_sid": call_sid,
            "message": "Simulated voice call initiated successfully",
            "phone_number": phone_number,
            "source": source,
            "simulated": True
        }
    
    def _simulate_conversation_flow(self, call_sid: str, phone_number: str):
        """Simulate the conversation flow"""
        logger.info(f"   🤖 AI: Hello! Welcome to Goodness Glamour Salon. I'm your AI assistant.")
        logger.info(f"   👤 Customer: Hi, I want to book an appointment")
        logger.info(f"   🤖 AI: Great! What's your name?")
        logger.info(f"   👤 Customer: My name is Sarah")
        logger.info(f"   🤖 AI: Nice to meet you, Sarah! Which service would you like?")
        logger.info(f"   👤 Customer: I want a haircut")
        logger.info(f"   🤖 AI: Perfect! What date works for you?")
        logger.info(f"   👤 Customer: Tomorrow")
        logger.info(f"   🤖 AI: Great! What time would work? We're available 9 AM to 8 PM.")
        logger.info(f"   👤 Customer: 2 PM")
        logger.info(f"   🤖 AI: Perfect! What's your address for our doorstep service?")
        logger.info(f"   👤 Customer: 123 Main Street, Mumbai")
        logger.info(f"   🤖 AI: Let me confirm your booking...")
        
        # Create a simulated booking
        booking_id = f"BG{int(time.time())}"
        logger.info(f"   🤖 AI: Booking confirmed! Booking ID: {booking_id}")
        logger.info(f"   🤖 AI: You'll receive an SMS confirmation shortly.")
        logger.info(f"   🤖 AI: Thank you for choosing Goodness Glamour!")
        
        # Simulate booking data
        simulated_booking = {
            "booking_id": booking_id,
            "customer_name": "Sarah",
            "phone": phone_number,
            "service": "haircut",
            "date": "tomorrow",
            "time": "2 PM",
            "address": "123 Main Street, Mumbai",
            "status": "confirmed",
            "created_at": datetime.now().isoformat(),
            "source": "simulated_voice_call"
        }
        
        # Save simulated booking
        self._save_booking_data(simulated_booking)
        
        logger.info(f"   ✅ Simulated booking saved: {booking_id}")
    
    def _save_booking_data(self, booking_data: dict):
        """Save booking data to Excel and database"""
        try:
            # Save to Excel
            self._save_to_excel(booking_data)
            
            # Save to SQLite
            self._save_to_database(booking_data)
            
            logger.info(f"Simulated booking saved: {booking_data['booking_id']}")
        except Exception as e:
            logger.error(f"Error saving simulated booking: {e}")

# Initialize the assistant
voice_assistant = SimpleVoiceAssistant()

# Flask app
app = Flask(__name__)

# QR Landing Page Template
QR_LANDING_PAGE_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Goodness Glamour - Voice Booking</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .container {
            background: white;
            border-radius: 20px;
            padding: 40px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            text-align: center;
            max-width: 400px;
            width: 100%;
        }
        .logo {
            font-size: 2em;
            font-weight: bold;
            color: #667eea;
            margin-bottom: 10px;
        }
        .phone-input {
            width: 100%;
            padding: 15px;
            border: 2px solid #ddd;
            border-radius: 10px;
            font-size: 16px;
            margin-bottom: 20px;
            box-sizing: border-box;
        }
        .call-button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 15px 30px;
            border-radius: 10px;
            font-size: 16px;
            font-weight: bold;
            cursor: pointer;
            width: 100%;
            margin-bottom: 20px;
        }
        .call-button:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
        }
        .status {
            margin-top: 20px;
            padding: 10px;
            border-radius: 5px;
            display: none;
        }
        .status.success {
            background: #d4edda;
            color: #155724;
        }
        .status.error {
            background: #f8d7da;
            color: #721c24;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="logo">💇‍♀️ Goodness Glamour</div>
        <h2>Book Your Appointment via Voice Call</h2>
        <p>Enter your phone number and we'll call you instantly!</p>
        
        <form id="bookingForm">
            <input type="tel" id="phoneInput" class="phone-input" 
                   placeholder="Enter your phone number (e.g., +919876543210)" required>
            <button type="submit" class="call-button" id="callButton">
                📞 Call Me Now
            </button>
        </form>
        
        <div id="status" class="status"></div>
    </div>

    <script>
        document.getElementById('bookingForm').addEventListener('submit', async function(e) {
            e.preventDefault();
            
            const phoneInput = document.getElementById('phoneInput');
            const callButton = document.getElementById('callButton');
            const status = document.getElementById('status');
            
            const phoneNumber = phoneInput.value.trim();
            
            if (!phoneNumber) {
                showStatus('Please enter your phone number', 'error');
                return;
            }
            
            callButton.disabled = true;
            callButton.textContent = 'Calling...';
            
            try {
                const response = await fetch('/trigger-voice-call', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json',
                    },
                    body: JSON.stringify({
                        phone_number: phoneNumber,
                        source: 'qr_code'
                    })
                });
                
                const result = await response.json();
                
                if (result.success) {
                    showStatus('Call initiated! We\'ll be calling you shortly...', 'success');
                    callButton.textContent = 'Call Initiated ✓';
                } else {
                    showStatus('Failed to initiate call. Please try again.', 'error');
                    callButton.disabled = false;
                    callButton.textContent = '📞 Call Me Now';
                }
            } catch (error) {
                showStatus('Network error. Please try again.', 'error');
                callButton.disabled = false;
                callButton.textContent = '📞 Call Me Now';
            }
        });
        
        function showStatus(message, type) {
            const status = document.getElementById('status');
            status.textContent = message;
            status.className = 'status ' + type;
            status.style.display = 'block';
        }
    </script>
</body>
</html>
"""

@app.route('/voice/incoming', methods=['POST'])
def handle_incoming_call():
    """Handle incoming voice call from QR code scan"""
    try:
        call_sid = request.form.get('CallSid')
        from_number = request.form.get('From')
        
        logger.info(f"Incoming call from {from_number}, SID: {call_sid}")
        
        # Start new session
        session = voice_assistant.start_session(call_sid, from_number)
        
        # Create TwiML response
        response = VoiceResponse()
        
        # Welcome message
        response.say(
            "Hello! Welcome to Goodness Glamour Salon. I'm your AI assistant. How can I help you today?",
            voice='Polly.Joanna',
            language='en-US'
        )
        
        # Gather user input
        gather = response.gather(
            input='speech',
            action=f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}',
            speech_timeout='auto',
            language='en-US',
            enhanced=True
        )
        
        # Fallback
        response.say("I didn't hear anything. Please speak after the tone.")
        response.redirect(f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}')
        
        return str(response)
        
    except Exception as e:
        logger.error(f"Error handling incoming call: {e}")
        response = VoiceResponse()
        response.say("I'm sorry, there was a technical issue. Please try again later.")
        response.hangup()
        return str(response)

@app.route('/voice/process_speech/<call_sid>', methods=['POST'])
def process_speech(call_sid):
    """Process speech input from Twilio"""
    try:
        speech_result = request.form.get('SpeechResult', '')
        
        if not speech_result:
            response = VoiceResponse()
            response.say("I didn't hear anything. Please speak clearly.")
            response.redirect(f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}')
            return str(response)
        
        logger.info(f"Processing speech for call {call_sid}: {speech_result}")
        
        # Get AI response
        ai_response = voice_assistant.get_conversation_response(speech_result, call_sid)
        
        # Create TwiML response
        response = VoiceResponse()
        response.say(ai_response, voice='Polly.Joanna', language='en-US')
        
        # Check if conversation is complete
        session = voice_assistant.active_sessions.get(call_sid)
        if session and session.current_step == "booking_complete":
            response.hangup()
            voice_assistant.end_session(call_sid)
        else:
            # Continue conversation
            gather = response.gather(
                input='speech',
                action=f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}',
                speech_timeout='auto',
                language='en-US',
                enhanced=True
            )
            
            response.say("I'm listening.")
            response.redirect(f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}')
        
        return str(response)
        
    except Exception as e:
        logger.error(f"Error processing speech: {e}")
        response = VoiceResponse()
        response.say("I'm sorry, I'm having trouble understanding. Please try again.")
        response.redirect(f'{WEBHOOK_BASE_URL}/voice/process_speech/{call_sid}')
        return str(response)

@app.route('/voice/status/<call_sid>', methods=['POST'])
def call_status(call_sid):
    """Handle call status updates"""
    call_status = request.form.get('CallStatus')
    logger.info(f"Call {call_sid} status: {call_status}")
    
    # Clean up if call ended
    if call_status in ['completed', 'busy', 'no-answer', 'failed', 'canceled']:
        voice_assistant.end_session(call_sid)
    
    return Response(status=200)

@app.route('/qr/voice-booking', methods=['GET'])
def qr_landing_page():
    """Landing page for QR code scans"""
    return render_template_string(QR_LANDING_PAGE_TEMPLATE)

@app.route('/trigger-voice-call', methods=['POST'])
def trigger_voice_call():
    """API endpoint to trigger voice call"""
    try:
        data = request.get_json()
        phone_number = data.get('phone_number')
        source = data.get('source', 'qr_code')
        
        if not phone_number:
            return jsonify({
                "success": False,
                "message": "Phone number is required"
            }), 400
        
        # Trigger voice call
        result = voice_assistant.trigger_voice_call(phone_number, source)
        
        if result["success"]:
            return jsonify(result), 200
        else:
            return jsonify(result), 500
            
    except Exception as e:
        logger.error(f"Error in trigger voice call endpoint: {e}")
        return jsonify({
            "success": False,
            "message": "Internal server error"
        }), 500

@app.route('/api/qr/generate', methods=['GET'])
def generate_qr_code():
    """Generate QR code for voice booking"""
    try:
        service_id = request.args.get('service_id')
        source = request.args.get('source', 'website')
        
        qr_result = voice_assistant.generate_qr_code(service_id, source)
        
        if qr_result:
            return jsonify(qr_result), 200
        else:
            return jsonify({
                "error": "Failed to generate QR code"
            }), 500
            
    except Exception as e:
        logger.error(f"Error generating QR code: {e}")
        return jsonify({
            "error": "Internal server error"
        }), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint with Twilio verification"""
    health_status = {
        "status": "healthy",
        "service": "Simple Voice Booking Assistant",
        "timestamp": datetime.now().isoformat(),
        "active_sessions": len(voice_assistant.active_sessions),
        "twilio_status": "unknown"
    }
    
    # Verify Twilio connectivity
    if client:
        try:
            account = client.api.accounts(TWILIO_ACCOUNT_SID).fetch()
            health_status["twilio_status"] = "connected"
            health_status["twilio_account"] = account.friendly_name
            health_status["twilio_balance"] = str(account.balance)
        except Exception as e:
            health_status["twilio_status"] = "error"
            health_status["twilio_error"] = str(e)
    else:
        health_status["twilio_status"] = "not_initialized"
    
    return jsonify(health_status)

@app.route('/verify-twilio', methods=['GET'])
def verify_twilio():
    """Verify Twilio account and return detailed information"""
    if not client:
        return jsonify({
            "success": False,
            "error": "Twilio client not initialized"
        }), 500
    
    try:
        # Get account information
        account = client.api.accounts(TWILIO_ACCOUNT_SID).fetch()
        
        # Get phone numbers
        incoming_phone_numbers = client.incoming_phone_numbers.list(limit=10)
        phone_numbers = []
        for number in incoming_phone_numbers:
            phone_numbers.append({
                "phone_number": number.phone_number,
                "friendly_name": number.friendly_name,
                "capabilities": number.capabilities
            })
        
        return jsonify({
            "success": True,
            "account": {
                "friendly_name": account.friendly_name,
                "status": account.status,
                "balance": str(account.balance),
                "type": account.type,
                "sid": account.sid
            },
            "phone_numbers": phone_numbers,
            "configured_phone": TWILIO_PHONE_NUMBER
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 7001))
    logger.info(f"Starting Simple AI Voice Booking Assistant on port {port}")
    app.run(host='0.0.0.0', port=port, debug=True)
